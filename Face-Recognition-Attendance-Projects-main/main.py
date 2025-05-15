# === main.py ===
import cv2
import numpy as np
import face_recognition
import os
import time
from datetime import datetime
import tkinter as tk
from tkinter import ttk
from threading import Thread
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import webbrowser




path = 'Training_images'
images = []
classNames = []

if not os.path.exists(path):
    os.makedirs(path)

for cl in os.listdir(path):
    img_path = os.path.join(path, cl)
    img = cv2.imread(img_path)
    if img is not None:
        images.append(img)
        classNames.append(os.path.splitext(cl)[0])
    else:
        print(f"Could not load image: {cl}")

def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)
        if encode:
            encodeList.append(encode[0])
        else:
            print("No face found in image.")
    return encodeList

encodeListKnown = findEncodings(images)





excel_file = 'Attendance.xlsx'
today = datetime.now().strftime('%Y-%m-%d')

def get_current_class_number():
    try:
        wb = load_workbook(excel_file)
        if today in wb.sheetnames:
            ws = wb[today]
            headers = [cell.value for cell in ws[1]]
            class_headers = [h for h in headers if h and h.startswith('Class ')]
            if class_headers:
                return int(class_headers[-1].split(' ')[1])
    except FileNotFoundError:
        pass
    return 1

class_number = get_current_class_number()

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = today
    ws.append(['Name', 'Date', 'Class 1'])
    wb.save(excel_file)
else:
    wb = load_workbook(excel_file)
    if today not in wb.sheetnames:
        ws = wb.create_sheet(title=today)
        headers = ['Name', 'Date']
        for i in range(1, class_number + 1):
            headers.append(f'Class {i}')
        ws.append(headers)
        wb.save(excel_file)





def markAttendance(name):
    global class_number
    now = datetime.now()
    time_string = now.strftime('%H:%M:%S')

    wb = load_workbook(excel_file)
    ws = wb[today]

    name_found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            headers = [cell.value for cell in ws[1]]
            for i, header in enumerate(headers):
                if header == f'Class {class_number}':
                    ws.cell(row=row[0].row, column=i + 1, value=time_string)
                    name_found = True
                    break
            break

    if not name_found:
        headers = [cell.value for cell in ws[1]]
        col_index = next((i + 1 for i, h in enumerate(headers) if h == f'Class {class_number}'), ws.max_column + 1)
        if col_index > len(headers):
            ws[get_column_letter(col_index) + '1'] = f'Class {class_number}'
        new_row = [name, today] + [''] * (col_index - 3) + [time_string]
        ws.append(new_row)

    wb.save(excel_file)




camera_running = False

def start_camera(index=0):
    global camera_running
    camera_running = True

    cap = cv2.VideoCapture(index)
    if not cap.read()[0]:
        print(f"[ERROR] Camera {index} not accessible")
        return

    presence_start = {}
    last_seen = {}
    marked_present = set()

    while camera_running:
        success, img = cap.read()
        if not success:
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        faces = face_recognition.face_locations(imgS)
        encodes = face_recognition.face_encodings(imgS, faces)

        now = time.time()

        for encodeFace, faceLoc in zip(encodes, faces):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                y1, x2, y2, x1 = [v * 4 for v in faceLoc]
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.putText(img, name, (x1, y2 - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)

                last_seen[name] = now
                if name not in presence_start:
                    presence_start[name] = now
                elif now - presence_start[name] > 20 and name not in marked_present:
                    markAttendance(name)
                    marked_present.add(name)
                    print(f"{name} marked present")

        for name in list(last_seen):
            if now - last_seen[name] > 300:
                presence_start.pop(name, None)
                last_seen.pop(name, None)
                marked_present.discard(name)
                print(f"{name} reset")

        cv2.imshow("Face Recognition", img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()





def view_attendance():
    top = tk.Toplevel()
    top.title(f"Attendance - {today}")
    tree = ttk.Treeview(top)
    tree.pack(fill=tk.BOTH, expand=True)

    wb = load_workbook(excel_file)
    ws = wb[today]

    headers = [cell.value for cell in ws[1]]
    tree["columns"] = headers
    tree["show"] = "headings"
    for header in headers:
        tree.heading(header, text=header)

    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)





def start_new_class():
    global class_number
    class_number += 1
    wb = load_workbook(excel_file)
    ws = wb[today]
    col = get_column_letter(ws.max_column + 1)
    ws[f"{col}1"] = f"Class {class_number}"
    wb.save(excel_file)
    print(f"Started Class {class_number}")

def quit_app(root):
    global camera_running
    camera_running = False
    time.sleep(1)
    cv2.destroyAllWindows()
    root.destroy()

def launch_gui():
    root = tk.Tk()
    root.title("Face Attendance System")
    root.geometry("400x400")

    tk.Label(root, text="Face Recognition Attendance", font=("Helvetica", 16)).pack(pady=20)

    cams = [str(i) for i in range(5) if cv2.VideoCapture(i).read()[0]]
    selected_cam = tk.StringVar(value=cams[0] if cams else "0")

    tk.Label(root, text="Select Camera:").pack()
    ttk.Combobox(root, textvariable=selected_cam, values=cams).pack(pady=5)

    tk.Button(root, text="Start Camera", bg="green", fg="white", command=lambda: Thread(target=start_camera, args=(int(selected_cam.get()),)).start()).pack(pady=10)
    tk.Button(root, text="Start New Class", bg="orange", fg="white", command=start_new_class).pack(pady=10)
    tk.Button(root, text="View Attendance", bg="blue", fg="white", command=view_attendance).pack(pady=10)
    tk.Button(root, text="Quit", bg="red", fg="white", command=lambda: quit_app(root)).pack(pady=10)

    root.mainloop()




webbrowser.open(f"file://{os.path.abspath('index.html')}")
launch_gui()
