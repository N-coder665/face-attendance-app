from flask import Flask, render_template, jsonify
from openpyxl import load_workbook
from datetime import datetime
import os


app = Flask(__name__)
excel_file = 'Attendance.xlsx'

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/attendance')
def get_attendance():
    today = datetime.now().strftime('%Y-%m-%d')
    if not os.path.exists(excel_file):
        return jsonify([])

    wb = load_workbook(excel_file)
    if today not in wb.sheetnames:
        return jsonify([])

    ws = wb[today]
    headers = [cell.value for cell in ws[1]]
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = {headers[i]: row[i] for i in range(len(headers))}
        data.append(entry)

    return jsonify(data)

if __name__ == '__main__':
    app.run(debug=True)
