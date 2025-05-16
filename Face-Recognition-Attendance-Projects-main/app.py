from flask import Flask, render_template, redirect, url_for
from openpyxl import load_workbook
from datetime import datetime
import subprocess

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/start_attendance')
def start_attendance():
    subprocess.Popen(["python", "main.py"])  # Launches the attendance system
    return redirect(url_for('index'))

@app.route('/attendance')
def attendance():
    today = datetime.now().strftime('%Y-%m-%d')
    try:
        wb = load_workbook("Attendance.xlsx")
        if today in wb.sheetnames:
            ws = wb[today]
            data = [list(row) for row in ws.iter_rows(values_only=True)]
            headers = data[0]
            rows = data[1:]
            return render_template("attendance.html", headers=headers, rows=rows)
        else:
            return render_template("attendance.html", headers=[], rows=[], no_data=True)
    except FileNotFoundError:
        return render_template("attendance.html", headers=[], rows=[], no_data=True)

if __name__ == '__main__':
    app.run(debug=True)